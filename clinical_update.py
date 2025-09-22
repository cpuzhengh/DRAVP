import time
import openpyxl as xl
from selenium.webdriver import Chrome, ChromeOptions
from selenium.common.exceptions import StaleElementReferenceException,NoSuchElementException
from selenium.webdriver.common.by import By

wb = xl.load_workbook('clinical_information.xlsx')
ws = wb.active
row = ws.max_row
print(row)

option = ChromeOptions()
option.add_argument('headless')
option.add_argument('blink-settings=imagesEnabled=false')

for i in range(2, row+1):
    NCT_number = ws.cell(row=i, column=2).value

    url = 'https://clinicaltrials.gov/ct2/show/' + str(NCT_number)

    web = Chrome(options=option)
    web.get(url)

    try:
        web.implicitly_wait(3)
        phase = web.find_element(By.XPATH, '//*[@id="tab-body"]/div/div[1]/div[2]/table/tbody/tr[2]/td[3]/span').text
    except NoSuchElementException:
        phase = 'Not available'

    print(i)
    print(phase)
    ws.cell(row=i, column=4).value = phase

    if i%20 == 0:
        wb.save('clinical_information.xlsx')

wb.save('clinical_information.xlsx')
wb.close()

